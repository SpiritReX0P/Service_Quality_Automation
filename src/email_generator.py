from openpyxl import load_workbook

from config.email_template import (
    EMAIL_GREETING,
    EMAIL_NOTE,
    PARAMETER_TABLE_HTML,
    EMAIL_SIGNATURE,
    DRIVE_LINK
)

# Import both KPI dictionaries
from config.kpi_thresholds import (
    EDC_KPI_RULES,
    ZONE_KPI_RULES
)


# =====================================================
# Detect Report Type
# =====================================================

def get_kpi_rules(ws):
    """
    Zone report has 15 columns.
    EDC report has 16 columns.
    """

    total_columns = ws.max_column

    if total_columns == 15:
        return ZONE_KPI_RULES

    return EDC_KPI_RULES


# =====================================================
# KPI COLOR LOGIC
# =====================================================

def get_kpi_style(col_idx, value, kpi_rules):

    if col_idx not in kpi_rules:
        return ""

    if value is None:
        return ""

    rule = kpi_rules[col_idx]

    try:

        # ==========================================
        # Higher is Better
        # ==========================================
        if rule["type"] == "higher":

            value = float(value) * 100

            if value >= rule["green"]:
                bg = "#d4edda"

            elif value >= rule["yellow"]:
                bg = "#fff3cd"

            else:
                bg = "#f8d7da"

        # ==========================================
        # Lower % is Better
        # ==========================================
        elif rule["type"] == "lower_pct":

            value = float(value) * 100

            if value <= rule["green"]:
                bg = "#d4edda"

            elif value <= rule["yellow"]:
                bg = "#fff3cd"

            else:
                bg = "#f8d7da"

        # ==========================================
        # Lower Count is Better
        # ==========================================
        else:

            value = float(value)

            if value <= rule["green"]:
                bg = "#d4edda"

            elif value <= rule["yellow"]:
                bg = "#fff3cd"

            else:
                bg = "#f8d7da"

        return f"""
        background-color:{bg};
        font-weight:bold;
        """

    except:
        return ""


# =====================================================
# EXCEL TO HTML TABLE
# =====================================================

def excel_to_html_table(report_file):

    wb = load_workbook(report_file)
    ws = wb.active

    # Detect whether this is Zone or EDC
    KPI_RULES = get_kpi_rules(ws)

    html = """
    <table
    style="
    border-collapse:collapse;
    font-family:Calibri;
    font-size:11pt;
    text-align:center;
    "
    border="1"
    cellpadding="4">
    """

    # =====================================================
    # STORE MERGED CELLS
    # =====================================================

    merged_cells = {}

    for merged_range in ws.merged_cells.ranges:

        min_col = merged_range.min_col
        max_col = merged_range.max_col
        min_row = merged_range.min_row
        max_row = merged_range.max_row

        merged_cells[(min_row, min_col)] = (
            max_row - min_row + 1,
            max_col - min_col + 1
        )

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):

                if (r, c) != (min_row, min_col):
                    merged_cells[(r, c)] = "SKIP"

    # =====================================================
    # BUILD HTML TABLE
    # =====================================================

    for row_idx, row in enumerate(ws.iter_rows(), start=1):

        html += "<tr>"

        for col_idx, cell in enumerate(row, start=1):

            if merged_cells.get((row_idx, col_idx)) == "SKIP":
                continue

            rowspan = 1
            colspan = 1

            if (row_idx, col_idx) in merged_cells:
                rowspan, colspan = merged_cells[(row_idx, col_idx)]

            value = cell.value
            raw_value = value

            if value is None:
                value = ""

            elif isinstance(value, (int, float)):

                if "%" in str(cell.number_format):
                    value = f"{value:.2%}"

                elif "#,##0" in str(cell.number_format):
                    value = f"{value:,.0f}"

            style = ""

            # =====================================================
            # Header Row 1
            # =====================================================

            if row_idx == 1:

                style = """
                background-color:#dfead8;
                font-weight:bold;
                """

            # =====================================================
            # Header Row 2
            # =====================================================

            elif row_idx == 2:

                style = """
                background-color:#d9e7f5;
                font-weight:bold;
                """

            # =====================================================
            # Total Row
            # =====================================================

            elif "Total" in str(row[0].value):

                style = """
                font-weight:bold;
                """

                style += get_kpi_style(
                    col_idx,
                    raw_value,
                    KPI_RULES
                )

            # =====================================================
            # Normal Data Rows
            # =====================================================

            else:

                style += get_kpi_style(
                    col_idx,
                    raw_value,
                    KPI_RULES
                )

            html += f"""
            <td
                rowspan="{rowspan}"
                colspan="{colspan}"
                style="{style}">
                {value}
            </td>
            """

        html += "</tr>"

    html += "</table>"

    return html


# =====================================================
# BUILD EMAIL BODY
# =====================================================

def build_email_body(report_file):

    report_table = excel_to_html_table(report_file)

    body = f"""
    <html>
    <body>

    {EMAIL_GREETING}

    <br><br>

    {report_table}

    <br><br>

    {EMAIL_NOTE}

    <br><br>

    {PARAMETER_TABLE_HTML}

    <br><br>

    PFA Drive link for Undelivered and Partial Delivery data:

    <br><br>

    <a href="{DRIVE_LINK}">
        Service_Quality_Improvement_Drive
    </a>

    <br><br>

    {EMAIL_SIGNATURE}

    </body>
    </html>
    """

    return body
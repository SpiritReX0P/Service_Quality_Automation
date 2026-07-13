from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import Border
from openpyxl.styles import Side
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from openpyxl import load_workbook

# ==========================================================
# COMMON STYLES
# PLACE HERE
# ==========================================================

bold_font = Font(
    bold=True,
    size=11
)

header1_fill = PatternFill(
    fill_type="solid",
    start_color="D9EAD3",
    end_color="D9EAD3"
)

header2_fill = PatternFill(
    fill_type="solid",
    start_color="BDD7EE",
    end_color="BDD7EE"
)

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)


#=============================================================================
# ZONE TEMPLATE BUILDER
#=============================================================================

# ==========================================================
# Purpose:
# Create Zone Report Template with:
# 1. Header Row 1 (Merged Headers)
# 2. Header Row 2 (Column Headers)
# 3. Formatting
# 4. Colors
# 5. Borders
# 6. Row Heights
# 7. Column Widths
# 8. Freeze Panes
# ==========================================================

# ==========================================================
# ZONE REPORT COLUMNS (ROW 2)
# ==========================================================

ZONE_COLUMNS = [
    "Zone",
    "EDC",
    "count",
    "%age",
    "Day Arrival",
    "GSL",
    "GNPS",
    "0-7",
    "8-15",
    "16-30",
    "31+",
    "7+%",
    "Previous DayTFD",
    "MTD TFD Conversion",
    "PU Conversion"
]

ZONE_WIDTHS = {
        "A": 15,   # Zone
        "B": 20,   # EDC
        "C": 12,   # Count
        "D": 12,   # %
        "E": 15,   # Day Arrival
        "F": 12,   # GSL
        "G": 12,   # GNPS
        "H": 10,   # 0-7
        "I": 10,   # 8-15
        "J": 10,   # 16-30
        "K": 10,   # 31+
        "L": 10,   # 7+%
        "M": 15,   # TFD
        "N": 18,   # Conversion
        "O": 18    # PU Conversion
    }

# ==========================================================
# CREATE ZONE TEMPLATE
# ==========================================================

def create_zone_template(file_path):

    # ======================================================
    # OPEN WORKBOOK
    # ======================================================

    wb = load_workbook(file_path)

    ws = wb.active

    # ======================================================
    # CLEAR EXISTING DATA (OPTIONAL)
    # ======================================================

    # ws.delete_rows(1, ws.max_row)

    # ======================================================
    # HEADER ROW 1 MERGES
    # ======================================================

    ws.merge_cells("C1:D1")
    ws.merge_cells("F1:G1")
    ws.merge_cells("H1:L1")
    ws.merge_cells("M1:N1")

    # ======================================================
    # HEADER ROW 1 VALUES
    # ======================================================

    ws["A1"] = "Zone"

    ws["B1"] = "EDC"

    ws["C1"] = "YTD Partial Deliveries"

    ws["E1"] = "MTD"

    ws["F1"] = "MTD GSL"

    ws["H1"] = "YTD Undelivered"

    ws["M1"] = "TFD"

    ws["O1"] = "MTD Pick Up"

    # ======================================================
    # HEADER ROW 2 VALUES
    # ======================================================

    for col_num, column_name in enumerate(
            ZONE_COLUMNS,
            start=1):

        ws.cell(
            row=2,
            column=col_num
        ).value = column_name


    # ======================================================
    # FORMAT HEADER ROW 1
    # ======================================================

    for cell in ws[1]:

        cell.font = bold_font

        cell.fill = header1_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border

    # ======================================================
    # FORMAT HEADER ROW 2
    # ======================================================

    for cell in ws[2]:

        cell.font = bold_font

        cell.fill = header2_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border

    # ======================================================
    # ROW HEIGHTS
    # ====================================================== 

    ws.row_dimensions[1].height = 25

    ws.row_dimensions[2].height = 22

    # ======================================================
    # COLUMN WIDTHS
    # ======================================================

    for col, width in ZONE_WIDTHS.items():

        ws.column_dimensions[col].width = width

    # ======================================================
    # FREEZE HEADER
    # ======================================================

    ws.freeze_panes = "A3"

    # ======================================================
    # SAVE FILE
    # ======================================================

    wb.save(file_path)

    print(
        f"Template created successfully -> {file_path}"
    )


#=============================================================================
# EXCEL TEMPLATE BUILDER
#=============================================================================

# ==========================================================
# Purpose:
# Create EDC Report Template with:
# 1. Header Row 1 (Merged Headers)
# 2. Header Row 2 (Column Headers)
# 3. Formatting
# 4. Colors
# 5. Borders
# 6. Row Heights
# 7. Column Widths
# 8. Freeze Panes
# ==========================================================

# ==========================================================
# EDC REPORT COLUMNS (ROW 2)
# ==========================================================

EDC_COLUMNS = [
    "Zone",
    "EDC",
    "GDW",
    "count",
    "%",
    "Day Arrival",
    "GSL",
    "GNPS",
    "0-7",
    "8-15",
    "16-30",
    "31+",
    "7+%",
    "Previous Day TFD",
    "MTD TFD Conversion",
    "PU Conversion"
]

EDC_WIDTHS = {
    "A": 15,
    "B": 20,
    "C": 12,
    "D": 12,
    "E": 15,
    "F": 12,
    "G": 12,
    "H": 10,
    "I": 10,
    "J": 10,
    "K": 10,
    "L": 10,
    "M": 25,
    "N": 25,
    "O": 22
}

# ==========================================================
# CREATE EDC TEMPLATE
# ==========================================================

def create_edc_template(file_path):

    # ======================================================
    # OPEN WORKBOOK
    # ======================================================

    wb = load_workbook(file_path)

    ws = wb.active

    # ======================================================
    # CLEAR EXISTING DATA (OPTIONAL)
    # ======================================================

    #ws.delete_rows(1, ws.max_row)

    # ======================================================
    # HEADER ROW 1 MERGES
    # ======================================================

    ws.merge_cells("D1:E1")
    ws.merge_cells("G1:H1")
    ws.merge_cells("I1:M1")
    ws.merge_cells("N1:O1")

    # ======================================================
    # HEADER ROW 1 VALUES
    # ======================================================

    ws["A1"] = "Zone"

    ws["B1"] = "EDC"

    ws["C1"] = "GDW"

    ws["D1"] = "YTD Partial Deliveries"

    ws["F1"] = "MTD"

    ws["G1"] = "MTD GSL"

    ws["I1"] = "YTD Undelivered"

    ws["N1"] = "TFD"

    ws["P1"] = "MTD Pick Up"
    

    # ======================================================
    # HEADER ROW 2 VALUES
    # ======================================================

    for col_num, value in enumerate(
        EDC_COLUMNS,
        start=1
    ):

        ws.cell(
            row=2,
            column=col_num
        ).value = value

    # ======================================================
    # FORMAT HEADER ROW 1
    # ======================================================

    for cell in ws[1]:

        cell.font = bold_font

        cell.fill = header1_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border

    # ======================================================
    # FORMAT HEADER ROW 2 
    # ======================================================

    for cell in ws[2]:

        cell.font = bold_font

        cell.fill = header2_fill

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        cell.border = thin_border

    # ======================================================
    # ROW HEIGHTS
    # ======================================================

    ws.row_dimensions[1].height = 25

    ws.row_dimensions[2].height = 22

    # ======================================================
    # COLUMN WIDTHS
    # ======================================================

    for col, width in EDC_WIDTHS.items():

        ws.column_dimensions[col].width = width

    # ======================================================
    # FREEZE HEADER
    # ======================================================

    ws.freeze_panes = "A3"

    # ======================================================
    # SAVE FILE
    # ======================================================
        
    wb.save(file_path)

    print(
        f"EDC Template Created -> {file_path}"
    )
import pandas as pd
from config.report_config import (
    ZONE_FILES,
    EDC_FILES
)
from openpyxl import load_workbook
from openpyxl.styles import Font


def load_master_workbook(master_file):

    edc_df = pd.read_excel(
        master_file,
        sheet_name="EDC",
        header=1
    )

    gdw_df = pd.read_excel(
        master_file,
        sheet_name="GDW",
        header=1
    )

    return edc_df, gdw_df


#===========================================
# FILTER DATA (ZONE'S,EDC'S)-> FUNCTION-1
#===========================================

#=====================
#FILTER DATA (ZONE'S)
#=====================
def get_zone_data(edc_df, zone_name):

    zone_df = edc_df[
        edc_df["Zone"].astype(str).str.contains(zone_name, na=False)
    ]

    return zone_df

#====================
#FILTER DATA (EDC'S)
#====================
def get_edc_data(gdw_df, edc_name):

    edc_df = gdw_df[
        gdw_df["EDC"].astype(str).str.contains(edc_name,na=False)
    ]

    return edc_df


#==================================================
# WRITE DATA TO EXCEL (ZONE'S,EDC'S)-> FUNCTION-2
#==================================================

#==============================
# WRITE DATA TO EXCEL (ZONE'S)
#==============================
def populate_zone_report(
        zone_df,
        report_file
):

    wb = load_workbook(report_file)

    ws = wb.active

    # Clear old report data
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    bold_font = Font(bold=True)

    start_row = 3

    for row_idx, (_, row) in enumerate(zone_df.iterrows(), start=start_row):

        ws.cell(row=row_idx, column=1).value = row["Zone"]
        ws.cell(row=row_idx, column=2).value = row["EDC"]
        ws.cell(row=row_idx, column=3).value = row["count"]
        cell = ws.cell(row=row_idx, column=4)
        cell.value = row["%age"]
        cell.number_format = "0.0000%"
        cell = ws.cell(row=row_idx, column=5)
        cell.value = row["Day Arrival"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=6)
        cell.value = row["GSL"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=7)
        cell.value = row["GNPS"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=8)
        cell.value = row["0-7"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=9)
        cell.value = row["8-15"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=10)
        cell.value = row["16-30"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=11)
        cell.value = row["31+"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=12)
        cell.value = row["7+"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=13)
        cell.value = row["Previous Day TFD"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=14)
        cell.value = row["MTD TFD Conversion"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=15)
        cell.value = row["PU Conversion"]
        cell.number_format = "0.00%"

        if "Total" in str(row["Zone"]):
            for col in range(1, 16):
                ws.cell(row=row_idx, column=col).font = bold_font

    wb.save(report_file)

    print(f"Data populated -> {report_file}")

#==============================
# WRITE DATA TO EXCEL (EDC'S)
#==============================
def populate_edc_report(
        edc_df,
        report_file
):

    wb = load_workbook(report_file)

    ws = wb.active

    # Clear old report data
    if ws.max_row >= 3:
        ws.delete_rows(3, ws.max_row - 2)

    bold_font = Font(bold=True)

    start_row = 3

    for row_idx, (_, row) in enumerate(
            edc_df.iterrows(),
            start=start_row
    ):

        ws.cell(row=row_idx, column=1).value = row["Zone"]
        ws.cell(row=row_idx, column=2).value = row["EDC"]
        ws.cell(row=row_idx, column=3).value = row["GDW"]
        cell = ws.cell(row=row_idx, column=4)
        cell.value = row["count"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=5)
        cell.value = row["%"]
        cell.number_format = "0.0000%"
        cell = ws.cell(row=row_idx, column=6)
        cell.value = row["Day Arrival"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=7)
        cell.value = row["GSL"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=8)
        cell.value = row["GNPS"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=9)
        cell.value = row["0-7"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=10)
        cell.value = row["8-15"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=11)
        cell.value = row["16-30"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=12)
        cell.value = row["31+"]
        cell.number_format = '#,##0'
        cell = ws.cell(row=row_idx, column=13)
        cell.value = row["7+"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=14)
        cell.value = row["TFD"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=15)
        cell.value = row["Conversion"]
        cell.number_format = "0.00%"
        cell = ws.cell(row=row_idx, column=16)
        cell.value = row["conversion"]
        cell.number_format = "0.00%"

        if "Total" in str(row["EDC"]):
            for col in range(1, 17):
                ws.cell(row=row_idx, column=col).font = bold_font
                
    wb.save(report_file)

    print(f"EDC populated -> {report_file}")    


#==========================================
# GENERATE ALL ZONE REPORTS
#==========================================

def populate_all_zone_reports(edc_df):

    print("\n===== GENERATING ZONE REPORTS =====")

    print(type(ZONE_FILES))
    print(ZONE_FILES)

    for zone_name, file_path in ZONE_FILES.items():

        print(f"Processing {zone_name}")

        filtered_df = get_zone_data(
            edc_df,
            zone_name
        )

        populate_zone_report(
            filtered_df,
            file_path
        )

    print("\nAll Zone Reports Generated Successfully")

#==========================================
# GENERATE ALL EDC REPORTS
#==========================================

def populate_all_edc_reports(gdw_df):

    print("\n===== GENERATING EDC REPORTS =====")

    for edc_name, file_path in EDC_FILES.items():

        print(f"Processing {edc_name}")

        filtered_df = get_edc_data(
            gdw_df,
            edc_name
        )

        populate_edc_report(
            filtered_df,
            file_path
        )

    print("\nAll EDC Reports Generated Successfully")